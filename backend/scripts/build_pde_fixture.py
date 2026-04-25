"""Build PdE test fixture (Sprint 3 prep).

Estrae ~40 righe rappresentative dal file PdE Numbers reale e le scrive
in un .xlsx committabile. Lo script è one-shot: l'output è la fixture
in `tests/fixtures/pde_sample.xlsx` (committato), non lo script stesso.

Lo script va rieseguito quando:
- Cambia il PdE source (nuovo anno) e si vogliono casi edge aggiornati
- Si scopre un nuovo pattern di periodicità non coperto dai test

Uso::

    PYTHONPATH=src uv run python scripts/build_pde_fixture.py \\
        --source "/path/to/PdE-YYYY.numbers"

Coverage cases (40 righe totali):
- 6 simple ("Circola tutti i giorni", no skip)
- 8 skip interval ("Non circola dal X al Y")
- 8 apply interval ("Circola dal X al Y")
- 6 date list lunga (>10 date)
- 8 date list corta (1-3 date)
- 2 doppia composizione SI
- 2 garantito festivo SI
"""

from __future__ import annotations

import argparse
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from numbers_parser import Document
from openpyxl import Workbook

DEFAULT_SOURCE = (
    "/Users/spant87/Library/Mobile Documents/com~apple~Numbers/Documents/"
    "All.1A5_14dic2025-12dic2026_TRENI e BUS_Rev5_RL.numbers"
)
DEFAULT_OUTPUT = Path("tests/fixtures/pde_sample.xlsx")


def categorize_rows(
    data_rows: list[list[Any]], col_period: int, col_doppia: int, col_gar_fest: int
) -> dict[str, list[int]]:
    """Suddivide gli indici per pattern di periodicità + flag."""
    buckets: dict[str, list[int]] = {
        "simple": [],
        "skip_interval": [],
        "apply_interval": [],
        "date_list_long": [],
        "date_list_short": [],
        "doppia_si": [],
        "garantito_festivo_si": [],
    }
    for i, row in enumerate(data_rows):
        p = row[col_period] or ""
        if not isinstance(p, str):
            continue
        pl = p.lower()
        slash_count = p.count("/")

        if "non circola dal" in pl:
            buckets["skip_interval"].append(i)
        elif "circola dal" in pl and "non circola" not in pl:
            buckets["apply_interval"].append(i)
        elif slash_count > 20:
            buckets["date_list_long"].append(i)
        elif slash_count > 0 and "circola" in pl and "tutti i giorni" not in pl and "dal" not in pl:
            buckets["date_list_short"].append(i)
        elif "tutti i giorni" in pl and "non circola" not in pl:
            buckets["simple"].append(i)

        if (row[col_doppia] or "") == "SI":
            buckets["doppia_si"].append(i)
        if (row[col_gar_fest] or "") == "SI":
            buckets["garantito_festivo_si"].append(i)
    return buckets


def select_fixture_indices(buckets: dict[str, list[int]]) -> list[int]:
    """Scelta deterministica: prendo i primi N indici per bucket."""
    selected: set[int] = set()
    quotas = {
        "simple": 6,
        "skip_interval": 8,
        "apply_interval": 8,
        "date_list_long": 6,
        "date_list_short": 8,
        "doppia_si": 2,
        "garantito_festivo_si": 2,
    }
    for bucket, n in quotas.items():
        for idx in buckets[bucket][:n]:
            selected.add(idx)
        # Se quota incompleta, ignoro (alcuni overlap tra bucket sono ok)
    return sorted(selected)


def write_xlsx(output: Path, header: list[Any], data: Iterable[list[Any]]) -> int:
    """Scrive un .xlsx con header + righe. Ritorna numero di righe dati."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl non ha creato un sheet di default")
    ws.title = "PdE RL"

    # Header
    for col_idx, val in enumerate(header, start=1):
        ws.cell(row=1, column=col_idx, value=val)

    # Data rows
    n = 0
    for row in data:
        n += 1
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=n + 1, column=col_idx, value=val)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return n


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source",
        default=DEFAULT_SOURCE,
        help=f"PdE Numbers file source (default: {DEFAULT_SOURCE})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output xlsx path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    print(f"Apertura PdE: {args.source}")
    doc = Document(args.source)
    table = doc.sheets[0].tables[0]
    rows = [[c.value for c in r] for r in table.rows()]
    header = rows[0]
    data_rows = rows[1:]
    print(f"  Header: {len(header)} colonne")
    print(f"  Data rows: {len(data_rows)}")

    col_period = header.index("Periodicità")
    col_doppia = header.index("Doppia Composizione - Invernale Feriale")
    col_gar_fest = header.index("Treno garantito festivo")

    buckets = categorize_rows(data_rows, col_period, col_doppia, col_gar_fest)
    print("  Bucket sizes:")
    for k, v in buckets.items():
        print(f"    {k:30s}: {len(v):5d}")

    indices = select_fixture_indices(buckets)
    print(f"\nIndici selezionati: {len(indices)}")
    print(f"  {indices}")

    selected_rows = [data_rows[i] for i in indices]
    n = write_xlsx(args.output, header, selected_rows)
    print(f"\nFixture scritta: {args.output} ({n} righe + header)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
