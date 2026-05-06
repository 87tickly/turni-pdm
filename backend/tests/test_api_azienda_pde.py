"""Test integration API ``/api/aziende/me/pde/*`` — Sub-MR 5.bis-d (entry 177).

Copre i 5 endpoint del PdE livello azienda:

- ``GET /pde/status`` (riepilogo)
- ``POST /pde/base`` (multipart upload PdE annuale)
- ``GET /variazioni`` (timeline globale)
- ``POST /variazioni`` (registra metadati globale)
- ``POST /variazioni/{run_id}/applica`` (multipart upload + applica)

Prefisso ``TEST_AZPDE_*`` per cleanup FK-safe.
"""

from __future__ import annotations

import io
import os
from collections.abc import AsyncGenerator
from datetime import date, time
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from colazione.db import dispose_engine, session_scope
from colazione.main import app

pytestmark = pytest.mark.skipif(
    os.getenv("SKIP_DB_TESTS") == "1",
    reason="DB not configured for tests",
)


# =====================================================================
# Fixture
# =====================================================================


_TRENO_PREFIX = "TAZP"
_STAZ_PREFIX = "STESTAZP"


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(app)


async def _wipe_data() -> None:
    """Cleanup FK-safe: corse → run → stazioni di test."""
    async with session_scope() as session:
        await session.execute(
            text("DELETE FROM corsa_commerciale WHERE numero_treno LIKE :p"),
            {"p": f"{_TRENO_PREFIX}%"},
        )
        await session.execute(
            text(
                "DELETE FROM corsa_import_run WHERE source_file LIKE :p"
            ),
            {"p": "TEST_AZPDE_%"},
        )
        # Le stazioni di test non sono ondelete CASCADE; le lascio (idempotenti).


@pytest.fixture(autouse=True)
async def _clean_per_test() -> AsyncGenerator[None, None]:
    await _wipe_data()
    yield


@pytest.fixture(scope="module", autouse=True)
async def _module_setup() -> AsyncGenerator[None, None]:
    yield
    await _wipe_data()
    await dispose_engine()


# =====================================================================
# Helpers
# =====================================================================


def _login(client: TestClient, username: str, password: str) -> str:
    res = client.post(
        "/api/auth/login",
        json={"username": username, "password": password},
    )
    assert res.status_code == 200, res.text
    return str(res.json()["access_token"])


def _giro_token(client: TestClient) -> str:
    return _login(client, "pianificatore_giro_demo", "demo12345")


def _admin_token(client: TestClient) -> str:
    return _login(client, "admin", "admin12345")


def _h(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def _crea_stazioni(codici: list[str]) -> None:
    """Inserisce stazioni di test (idempotente)."""
    async with session_scope() as session:
        for c in codici:
            await session.execute(
                text(
                    "INSERT INTO stazione (codice, nome, azienda_id) "
                    "SELECT :c, :n, a.id FROM azienda a WHERE a.codice = 'trenord' "
                    "ON CONFLICT (codice) DO NOTHING"
                ),
                {"c": c, "n": f"TEST {c}"},
            )
        await session.commit()


def _build_xlsx_pde_mini(rows: list[dict[str, Any]]) -> bytes:
    """xlsx PdE minimal-but-valid in-memory (replica del helper in
    ``test_api_variazioni_applica.py``)."""
    if not rows:
        raise ValueError("almeno 1 row necessaria")
    header_keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row:
            if k not in seen:
                header_keys.append(k)
                seen.add(k)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "PdE RL"
    for col_idx, key in enumerate(header_keys, start=1):
        ws.cell(row=1, column=col_idx, value=key)
    for row_idx, raw in enumerate(rows, start=2):
        for col_idx, key in enumerate(header_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=raw.get(key))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_pde_minimal(
    *,
    treno: str,
    cod_origine: str,
    cod_destinazione: str,
    valido_da: date,
    valido_a: date,
    ora_partenza: time = time(6, 39),
    ora_arrivo: time = time(7, 30),
) -> dict[str, Any]:
    return {
        "Treno 1": treno,
        "Cod Origine": cod_origine,
        "Cod Destinazione": cod_destinazione,
        "Valido da": valido_da,
        "Valido a": valido_a,
        "Ora Or": ora_partenza,
        "Ora Des": ora_arrivo,
        "Periodicità": "Circola tutti i giorni.",
        "Modalità di effettuazione": "T",
    }


# =====================================================================
# Test: GET /pde/status
# =====================================================================


async def test_status_401_senza_token(client: TestClient) -> None:
    res = client.get("/api/aziende/me/pde/status")
    assert res.status_code == 401


async def test_status_db_vuoto(client: TestClient) -> None:
    """Status su azienda senza PdE caricato: counters a zero, base_run None."""
    res = client.get(
        "/api/aziende/me/pde/status", headers=_h(_giro_token(client))
    )
    assert res.status_code == 200, res.text
    body = res.json()
    # Potrebbe esserci un base_run pre-esistente nel DB di test (il PdE
    # reale dell'azienda Trenord). Accetto quel caso ma verifico schema.
    assert "base_run" in body
    assert "n_corse_attive" in body
    assert "n_corse_totali" in body
    assert "n_variazioni_totali" in body
    assert "n_variazioni_applicate" in body
    assert isinstance(body["n_corse_attive"], int)
    assert isinstance(body["n_corse_totali"], int)
    assert body["n_corse_attive"] <= body["n_corse_totali"]


# =====================================================================
# Test: POST /pde/base (multipart upload)
# =====================================================================


async def test_carica_pde_base_415_estensione_invalida(
    client: TestClient,
) -> None:
    """File con estensione non supportata → 415."""
    res = client.post(
        "/api/aziende/me/pde/base",
        files={"file": ("test.txt", b"not a pde", "text/plain")},
        headers=_h(_giro_token(client)),
    )
    assert res.status_code == 415


async def test_carica_pde_base_xlsx_minimo_ok(client: TestClient) -> None:
    """Carica xlsx PdE mini con 1 corsa → 1 corsa creata in DB."""
    await _crea_stazioni(
        [f"{_STAZ_PREFIX}A", f"{_STAZ_PREFIX}B"]
    )
    xlsx_bytes = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}001",
                cod_origine=f"{_STAZ_PREFIX}A",
                cod_destinazione=f"{_STAZ_PREFIX}B",
                valido_da=date(2026, 1, 1),
                valido_a=date(2026, 1, 7),
            )
        ]
    )
    res = client.post(
        "/api/aziende/me/pde/base",
        files={
            "file": (
                "TEST_AZPDE_minimo.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_h(_giro_token(client)),
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["skipped"] is False
    assert body["run_id"] is not None
    assert body["n_create"] == 1
    assert body["n_total"] >= 1


async def test_carica_pde_base_idempotente_secondo_upload(
    client: TestClient,
) -> None:
    """Re-upload dello stesso file → skipped=True (idempotenza SHA-256)."""
    await _crea_stazioni([f"{_STAZ_PREFIX}A", f"{_STAZ_PREFIX}B"])
    xlsx_bytes = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}IDEM",
                cod_origine=f"{_STAZ_PREFIX}A",
                cod_destinazione=f"{_STAZ_PREFIX}B",
                valido_da=date(2026, 2, 1),
                valido_a=date(2026, 2, 7),
            )
        ]
    )
    files = {
        "file": (
            "TEST_AZPDE_idem.xlsx",
            xlsx_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    res1 = client.post(
        "/api/aziende/me/pde/base",
        files=files,
        headers=_h(_giro_token(client)),
    )
    assert res1.status_code == 200, res1.text
    assert res1.json()["skipped"] is False

    # Re-upload identico
    res2 = client.post(
        "/api/aziende/me/pde/base",
        files={
            "file": (
                "TEST_AZPDE_idem.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_h(_giro_token(client)),
    )
    assert res2.status_code == 200, res2.text
    body2 = res2.json()
    assert body2["skipped"] is True
    assert "skip_reason" in body2


# =====================================================================
# Test: POST /variazioni (registra metadati)
# =====================================================================


async def test_registra_variazione_globale_ok(client: TestClient) -> None:
    res = client.post(
        "/api/aziende/me/variazioni",
        json={
            "tipo": "INTEGRAZIONE",
            "source_file": "TEST_AZPDE_var.xlsx",
            "n_corse": 5,
            "note": "test entry 177",
        },
        headers=_h(_giro_token(client)),
    )
    assert res.status_code == 201, res.text
    body = res.json()
    assert body["tipo"] == "INTEGRAZIONE"
    assert body["programma_materiale_id"] is None
    assert body["completed_at"] is None


async def test_registra_variazione_globale_tipo_base_422(
    client: TestClient,
) -> None:
    """``BASE`` non ammesso: il caricamento BASE va via /pde/base."""
    res = client.post(
        "/api/aziende/me/variazioni",
        json={"tipo": "BASE", "source_file": "x.xlsx"},
        headers=_h(_giro_token(client)),
    )
    assert res.status_code == 422


async def test_registra_variazione_globale_403_admin_bypassa(
    client: TestClient,
) -> None:
    """L'admin ha is_admin=True e bypassa il check di ruolo."""
    res = client.post(
        "/api/aziende/me/variazioni",
        json={
            "tipo": "VARIAZIONE_INTERRUZIONE",
            "source_file": "TEST_AZPDE_admin.xlsx",
        },
        headers=_h(_admin_token(client)),
    )
    assert res.status_code == 201, res.text


# =====================================================================
# Test: GET /variazioni (lista globale)
# =====================================================================


async def test_list_variazioni_globali_ordinata_desc(
    client: TestClient,
) -> None:
    """Crea 2 variazioni → list ordinata DESC per started_at."""
    for tipo in ("INTEGRAZIONE", "VARIAZIONE_INTERRUZIONE"):
        client.post(
            "/api/aziende/me/variazioni",
            json={
                "tipo": tipo,
                "source_file": f"TEST_AZPDE_list_{tipo}.xlsx",
            },
            headers=_h(_giro_token(client)),
        )
    res = client.get(
        "/api/aziende/me/variazioni",
        headers=_h(_giro_token(client)),
    )
    assert res.status_code == 200, res.text
    body = res.json()
    nostre = [
        r for r in body if r["source_file"].startswith("TEST_AZPDE_list_")
    ]
    assert len(nostre) == 2
    # Tutte globali
    assert all(r["programma_materiale_id"] is None for r in nostre)
    # Tutte non-BASE
    assert all(r["tipo"] != "BASE" for r in nostre)


# =====================================================================
# Test: POST /variazioni/{run_id}/applica
# =====================================================================


async def test_applica_variazione_globale_run_inesistente_404(
    client: TestClient,
) -> None:
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}X",
                cod_origine=f"{_STAZ_PREFIX}A",
                cod_destinazione=f"{_STAZ_PREFIX}B",
                valido_da=date(2026, 5, 1),
                valido_a=date(2026, 5, 7),
            )
        ]
    )
    res = client.post(
        "/api/aziende/me/variazioni/9999999/applica",
        files={
            "file": (
                "TEST_AZPDE_apply404.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_h(_giro_token(client)),
    )
    assert res.status_code == 404


async def test_applica_variazione_globale_integrazione_smoke(
    client: TestClient,
) -> None:
    """Smoke: registra INTEGRAZIONE + applica con xlsx mini → 1 corsa
    inserita + run.completed_at popolato."""
    await _crea_stazioni([f"{_STAZ_PREFIX}A", f"{_STAZ_PREFIX}B"])

    # 1) Registra
    res_reg = client.post(
        "/api/aziende/me/variazioni",
        json={
            "tipo": "INTEGRAZIONE",
            "source_file": "TEST_AZPDE_smoke.xlsx",
        },
        headers=_h(_giro_token(client)),
    )
    assert res_reg.status_code == 201, res_reg.text
    run_id = res_reg.json()["id"]

    # 2) Applica
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}SMK",
                cod_origine=f"{_STAZ_PREFIX}A",
                cod_destinazione=f"{_STAZ_PREFIX}B",
                valido_da=date(2026, 8, 1),
                valido_a=date(2026, 8, 7),
            )
        ]
    )
    res = client.post(
        f"/api/aziende/me/variazioni/{run_id}/applica",
        files={
            "file": (
                "TEST_AZPDE_smoke.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_h(_giro_token(client)),
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["n_corse_create"] == 1
    assert body["completed_at"] is not None

    # 3) Re-applica → 409
    res_again = client.post(
        f"/api/aziende/me/variazioni/{run_id}/applica",
        files={
            "file": (
                "TEST_AZPDE_smoke.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_h(_giro_token(client)),
    )
    assert res_again.status_code == 409
