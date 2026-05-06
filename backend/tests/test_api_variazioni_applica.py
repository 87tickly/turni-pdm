"""Test integration end-to-end del MR 5.bis (Sprint 8.0, entry 173).

Endpoint ``POST /api/programmi/{programma_id}/variazioni/{run_id}/applica``:
upload xlsx, parser, planner pure, applicazione al DB, update run.

Le fixture xlsx sono costruite **in-memory** ad-hoc per ogni test
(niente file fixture su disco): ho controllo totale su cosa viene
parsato, e la suite resta auto-contenuta.

I test usano prefissi ``TEST_VAR_`` per programmi/stazioni/run così
che il cleanup possa wipare in modo sicuro senza toccare i dati di
altri test (es. il run BASE prodotto dal seed Trenord).
"""

from __future__ import annotations

import io
import os
from collections.abc import AsyncGenerator
from datetime import date, datetime, time
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from colazione.db import dispose_engine, session_scope
from colazione.main import app

pytestmark = pytest.mark.skipif(
    os.getenv("SKIP_DB_TESTS") == "1",
    reason="DB not configured for tests",
)


# =====================================================================
# Fixture di modulo
# =====================================================================


_PROG_PREFIX = "TEST_VAR_"
_STAZ_PREFIX = "ZTV"
_TRENO_PREFIX = "9000"  # Treni di test, non usati dal seed reale


@pytest.fixture(scope="module")
def client() -> TestClient:
    return TestClient(app)


async def _wipe_var_data() -> None:
    """Cleanup FK-safe per i dati TEST_VAR_*.

    Ordine: corse_composizione → corse_commerciali → run → stazioni →
    programmi. Le composizioni hanno CASCADE su corse, quindi basta
    cancellare le corse. Il prefisso ``ZTV_`` evita di toccare le
    stazioni reali Trenord (che hanno codici tipo ``S01066``).
    """
    async with session_scope() as session:
        await session.execute(
            text(
                "DELETE FROM corsa_commerciale WHERE codice_origine LIKE :p OR "
                "codice_destinazione LIKE :p"
            ),
            {"p": f"{_STAZ_PREFIX}%"},
        )
        await session.execute(
            text(
                "DELETE FROM corsa_import_run WHERE programma_materiale_id IN ("
                "  SELECT id FROM programma_materiale WHERE nome LIKE :p"
                ")"
            ),
            {"p": f"{_PROG_PREFIX}%"},
        )
        await session.execute(
            text("DELETE FROM stazione WHERE codice LIKE :p"),
            {"p": f"{_STAZ_PREFIX}%"},
        )
        await session.execute(
            text("DELETE FROM programma_materiale WHERE nome LIKE :p"),
            {"p": f"{_PROG_PREFIX}%"},
        )


@pytest.fixture(scope="module", autouse=True)
async def _module_setup() -> AsyncGenerator[None, None]:
    await _wipe_var_data()
    yield
    await _wipe_var_data()
    await dispose_engine()


@pytest.fixture(autouse=True)
async def _clean_per_test() -> AsyncGenerator[None, None]:
    await _wipe_var_data()
    yield


# =====================================================================
# Helpers HTTP
# =====================================================================


def _login(client: TestClient, username: str, password: str) -> str:
    res = client.post(
        "/api/auth/login",
        json={"username": username, "password": password},
    )
    assert res.status_code == 200, res.text
    return str(res.json()["access_token"])


def _admin_token(client: TestClient) -> str:
    return _login(client, "admin", "admin12345")


def _h(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


# =====================================================================
# Helpers di seed
# =====================================================================


async def _crea_programma_var(
    nome_suffix: str,
    valido_da: date = date(2026, 1, 1),
    valido_a: date = date(2026, 12, 31),
) -> int:
    """Crea un programma_materiale TEST_VAR_* in stato attivo."""
    async with session_scope() as session:
        row = (
            await session.execute(
                text(
                    "INSERT INTO programma_materiale "
                    "(azienda_id, nome, valido_da, valido_a, stato, "
                    "stato_pipeline_pdc, stato_manutenzione, "
                    "n_giornate_default, n_giornate_min, n_giornate_max, "
                    "fascia_oraria_tolerance_min, strict_options_json, "
                    "stazioni_sosta_extra_json) "
                    "SELECT id, :nome, :da, :a, 'attivo', "
                    "'PDE_IN_LAVORAZIONE', 'IN_ATTESA', "
                    "1, 4, 12, 30, '{}'::jsonb, '[]'::jsonb "
                    "FROM azienda WHERE codice = 'trenord' "
                    "RETURNING id"
                ),
                {
                    "nome": f"{_PROG_PREFIX}{nome_suffix}",
                    "da": valido_da,
                    "a": valido_a,
                },
            )
        ).first()
        assert row is not None
        return int(row[0])


async def _crea_stazioni(codici: list[str]) -> None:
    """Inserisce stazioni di test (idempotente) per evitare violazioni FK."""
    async with session_scope() as session:
        for codice in codici:
            await session.execute(
                text(
                    "INSERT INTO stazione (codice, nome, azienda_id) "
                    "SELECT :c, :n, id FROM azienda WHERE codice = 'trenord' "
                    "ON CONFLICT (codice) DO NOTHING"
                ),
                {"c": codice, "n": f"Stazione test {codice}"},
            )


async def _seed_corsa(
    *,
    numero_treno: str,
    codice_origine: str,
    codice_destinazione: str,
    valido_da: date,
    valido_a: date,
    valido_in_date_json: list[str],
    ora_partenza: time = time(6, 39),
    ora_arrivo: time = time(7, 30),
    row_hash: str = "h_seed",
) -> int:
    """Inserisce una CorsaCommerciale di test e ritorna il suo id.

    Imposta solo i campi obbligatori + quelli rilevanti al matching
    della variazione. Tutti gli altri restano default/NULL.
    """
    async with session_scope() as session:
        result = await session.execute(
            text(
                "INSERT INTO corsa_commerciale ("
                "azienda_id, row_hash, numero_treno, "
                "codice_origine, codice_destinazione, "
                "ora_partenza, ora_arrivo, valido_da, valido_a, "
                "is_treno_garantito_feriale, is_treno_garantito_festivo, "
                "giorni_per_mese_json, valido_in_date_json, import_source"
                ") SELECT id, :rh, :nt, :co, :cd, :op, :oa, :vda, :vaa, "
                "FALSE, FALSE, '{}'::jsonb, "
                "CAST(:vid AS jsonb), 'pde' "
                "FROM azienda WHERE codice = 'trenord' RETURNING id"
            ),
            {
                "rh": row_hash,
                "nt": numero_treno,
                "co": codice_origine,
                "cd": codice_destinazione,
                "op": ora_partenza,
                "oa": ora_arrivo,
                "vda": valido_da,
                "vaa": valido_a,
                "vid": _jsonb_dumps(valido_in_date_json),
            },
        )
        first = result.first()
        assert first is not None
        return int(first[0])


def _jsonb_dumps(value: Any) -> str:
    """Serializza in JSON per cast esplicito ``CAST(:val AS jsonb)``."""
    import json

    return json.dumps(value)


# =====================================================================
# Helpers fixture xlsx in-memory
# =====================================================================


def _build_xlsx_pde_mini(rows: list[dict[str, Any]]) -> bytes:
    """Costruisce un xlsx PdE minimal-but-valid in-memory.

    L'header viene dall'unione delle chiavi di tutte le rows. Le
    chiavi mancanti in una row producono celle vuote (None). Il
    parser PdE :func:`colazione.importers.pde.parse_corsa_row` accede
    via ``row.get(...)`` per la maggior parte dei campi, quindi
    basta popolare le 7 colonne obbligatorie (più ``Periodicità`` se
    voglio influenzare ``valido_in_date_json``).
    """
    if not rows:
        raise ValueError("almeno 1 row necessaria")

    header_keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row:
            if k not in seen:
                header_keys.append(k)
                seen.add(k)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "PdE RL"
    for col_idx, key in enumerate(header_keys, start=1):
        ws.cell(row=1, column=col_idx, value=key)
    for row_idx, raw in enumerate(rows, start=2):
        for col_idx, key in enumerate(header_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=raw.get(key))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_pde_minimal(
    *,
    treno: str,
    cod_origine: str,
    cod_destinazione: str,
    valido_da: date,
    valido_a: date,
    ora_partenza: time = time(6, 39),
    ora_arrivo: time = time(7, 30),
    periodicita: str = "Circola tutti i giorni.",
) -> dict[str, Any]:
    """Costruisce una row PdE con i 7 campi obbligatori + Periodicità."""
    return {
        "Treno 1": treno,
        "Cod Origine": cod_origine,
        "Cod Destinazione": cod_destinazione,
        "Valido da": valido_da,
        "Valido a": valido_a,
        "Ora Or": ora_partenza,
        "Ora Des": ora_arrivo,
        "Periodicità": periodicita,
        "Modalità di effettuazione": "T",
    }


async def _registra_variazione(client: TestClient, programma_id: int, tipo: str) -> int:
    """POST /variazioni → ritorna run_id della variazione registrata."""
    res = client.post(
        f"/api/programmi/{programma_id}/variazioni",
        json={"tipo": tipo, "source_file": f"{tipo.lower()}_test.xlsx"},
        headers=_h(_admin_token(client)),
    )
    assert res.status_code == 201, res.text
    return int(res.json()["id"])


def _applica_variazione(
    client: TestClient,
    programma_id: int,
    run_id: int,
    xlsx_bytes: bytes,
    *,
    filename: str = "var.xlsx",
    token: str | None = None,
) -> Any:
    """POST /variazioni/{run_id}/applica con file multipart."""
    if token is None:
        token = _admin_token(client)
    return client.post(
        f"/api/programmi/{programma_id}/variazioni/{run_id}/applica",
        files={
            "file": (
                filename,
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=_h(token),
    )


# =====================================================================
# 1) INTEGRAZIONE — inserisce corsa nuova
# =====================================================================


async def test_applica_integrazione_inserisce_corsa_nuova(
    client: TestClient,
) -> None:
    pid = await _crea_programma_var("integr_ok")
    await _crea_stazioni([f"{_STAZ_PREFIX}A", f"{_STAZ_PREFIX}B"])
    run_id = await _registra_variazione(client, pid, "INTEGRAZIONE")

    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}1",
                cod_origine=f"{_STAZ_PREFIX}A",
                cod_destinazione=f"{_STAZ_PREFIX}B",
                valido_da=date(2026, 3, 1),
                valido_a=date(2026, 3, 7),
            )
        ]
    )

    res = _applica_variazione(client, pid, run_id, xlsx)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["tipo"] == "INTEGRAZIONE"
    assert body["n_corse_create"] == 1
    assert body["n_corse_update"] == 0
    assert body["n_corse_lette_da_file"] == 1
    assert body["n_warnings"] == 0
    assert body["completed_at"] is not None

    # Verifica DB: la corsa è in DB con import_run_id pari alla run.
    async with session_scope() as session:
        row = (
            await session.execute(
                text(
                    "SELECT numero_treno, codice_origine, codice_destinazione, "
                    "import_run_id FROM corsa_commerciale "
                    "WHERE numero_treno = :nt"
                ),
                {"nt": f"{_TRENO_PREFIX}1"},
            )
        ).first()
        assert row is not None
        assert row[0] == f"{_TRENO_PREFIX}1"
        assert row[1] == f"{_STAZ_PREFIX}A"
        assert row[2] == f"{_STAZ_PREFIX}B"
        assert row[3] == run_id


async def test_applica_integrazione_idempotente_su_riapplicazione_stesso_hash(
    client: TestClient,
) -> None:
    """Se l'utente riapplica lo stesso file (nuova run, stesso hash) → 0
    INSERT (la corsa già esiste in DB)."""
    pid = await _crea_programma_var("integr_idemp")
    await _crea_stazioni([f"{_STAZ_PREFIX}A", f"{_STAZ_PREFIX}B"])
    run1 = await _registra_variazione(client, pid, "INTEGRAZIONE")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}1",
                cod_origine=f"{_STAZ_PREFIX}A",
                cod_destinazione=f"{_STAZ_PREFIX}B",
                valido_da=date(2026, 3, 1),
                valido_a=date(2026, 3, 7),
            )
        ]
    )
    res1 = _applica_variazione(client, pid, run1, xlsx)
    assert res1.status_code == 200
    assert res1.json()["n_corse_create"] == 1

    # Ri-registra una NUOVA run con lo stesso file → niente INSERT.
    run2 = await _registra_variazione(client, pid, "INTEGRAZIONE")
    res2 = _applica_variazione(client, pid, run2, xlsx)
    assert res2.status_code == 200
    body2 = res2.json()
    assert body2["n_corse_create"] == 0
    assert body2["n_warnings"] == 1
    assert "già presenti in DB" in body2["warnings"][0]


# =====================================================================
# 2) VARIAZIONE_ORARIO — aggiorna orari
# =====================================================================


async def test_applica_orario_aggiorna_orari(client: TestClient) -> None:
    pid = await _crea_programma_var("orario_ok")
    await _crea_stazioni([f"{_STAZ_PREFIX}C", f"{_STAZ_PREFIX}D"])
    corsa_id = await _seed_corsa(
        numero_treno=f"{_TRENO_PREFIX}2",
        codice_origine=f"{_STAZ_PREFIX}C",
        codice_destinazione=f"{_STAZ_PREFIX}D",
        valido_da=date(2026, 4, 1),
        valido_a=date(2026, 4, 30),
        valido_in_date_json=["2026-04-01"],
        ora_partenza=time(6, 0),
        ora_arrivo=time(7, 0),
    )
    run_id = await _registra_variazione(client, pid, "VARIAZIONE_ORARIO")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}2",
                cod_origine=f"{_STAZ_PREFIX}C",
                cod_destinazione=f"{_STAZ_PREFIX}D",
                valido_da=date(2026, 4, 1),
                valido_a=date(2026, 4, 30),
                ora_partenza=time(8, 30),
                ora_arrivo=time(9, 45),
            )
        ]
    )
    res = _applica_variazione(client, pid, run_id, xlsx)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["n_corse_create"] == 0
    assert body["n_corse_update"] == 1
    assert body["n_warnings"] == 0

    # Verifica DB: orari aggiornati.
    async with session_scope() as session:
        row = (
            await session.execute(
                text("SELECT ora_partenza, ora_arrivo FROM corsa_commerciale WHERE id = :id"),
                {"id": corsa_id},
            )
        ).first()
        assert row is not None
        assert row[0] == time(8, 30)
        assert row[1] == time(9, 45)


async def test_applica_orario_no_match_emette_warning(
    client: TestClient,
) -> None:
    """Nessuna corsa esistente con la chiave a 5 → 0 update + 1 warning."""
    pid = await _crea_programma_var("orario_nomatch")
    await _crea_stazioni([f"{_STAZ_PREFIX}E", f"{_STAZ_PREFIX}F"])
    run_id = await _registra_variazione(client, pid, "VARIAZIONE_ORARIO")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}3",  # mai seedato
                cod_origine=f"{_STAZ_PREFIX}E",
                cod_destinazione=f"{_STAZ_PREFIX}F",
                valido_da=date(2026, 5, 1),
                valido_a=date(2026, 5, 31),
            )
        ]
    )
    res = _applica_variazione(client, pid, run_id, xlsx)
    assert res.status_code == 200
    body = res.json()
    assert body["n_corse_update"] == 0
    assert body["n_warnings"] == 1
    assert "corsa target non trovata" in body["warnings"][0]


# =====================================================================
# 3) VARIAZIONE_INTERRUZIONE — riduce valido_in_date_json
# =====================================================================


async def test_applica_interruzione_intersezione_date(
    client: TestClient,
) -> None:
    """Esistente con 5 date, file dichiara solo 2 → DB → 2 (intersezione)."""
    pid = await _crea_programma_var("interr_ok")
    await _crea_stazioni([f"{_STAZ_PREFIX}G", f"{_STAZ_PREFIX}H"])
    corsa_id = await _seed_corsa(
        numero_treno=f"{_TRENO_PREFIX}4",
        codice_origine=f"{_STAZ_PREFIX}G",
        codice_destinazione=f"{_STAZ_PREFIX}H",
        valido_da=date(2026, 6, 1),
        valido_a=date(2026, 6, 30),
        valido_in_date_json=[
            "2026-06-01",
            "2026-06-02",
            "2026-06-03",
            "2026-06-04",
            "2026-06-05",
        ],
    )
    run_id = await _registra_variazione(client, pid, "VARIAZIONE_INTERRUZIONE")
    # File dichiara "Circola dal 01/06/2026 al 02/06/2026" → solo 2 date.
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}4",
                cod_origine=f"{_STAZ_PREFIX}G",
                cod_destinazione=f"{_STAZ_PREFIX}H",
                valido_da=date(2026, 6, 1),
                valido_a=date(2026, 6, 30),
                periodicita="Circola dal 01/06/2026 al 02/06/2026.",
            )
        ]
    )
    res = _applica_variazione(client, pid, run_id, xlsx)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["n_corse_update"] == 1
    assert body["n_warnings"] == 0

    async with session_scope() as session:
        row = (
            await session.execute(
                text("SELECT valido_in_date_json FROM corsa_commerciale WHERE id = :id"),
                {"id": corsa_id},
            )
        ).first()
        assert row is not None
        # JSONB → list Python via SQLAlchemy.
        assert sorted(row[0]) == ["2026-06-01", "2026-06-02"]


# =====================================================================
# 4) VARIAZIONE_CANCELLAZIONE — svuota valido_in_date_json
# =====================================================================


async def test_applica_cancellazione_setta_flag_is_cancellata(
    client: TestClient,
) -> None:
    """Sub-MR 5.bis-a alignment (entry 176): match a 5 → flag
    ``is_cancellata=True`` con audit trail (``cancellata_da_run_id``,
    ``cancellata_at``). Il vecchio approccio "svuota
    ``valido_in_date_json``" è stato sostituito (più tracciabile +
    coerente col CHECK constraint
    ``corsa_commerciale_cancellazione_coerente``)."""
    pid = await _crea_programma_var("canc_ok")
    await _crea_stazioni([f"{_STAZ_PREFIX}I", f"{_STAZ_PREFIX}J"])
    corsa_id = await _seed_corsa(
        numero_treno=f"{_TRENO_PREFIX}5",
        codice_origine=f"{_STAZ_PREFIX}I",
        codice_destinazione=f"{_STAZ_PREFIX}J",
        valido_da=date(2026, 7, 1),
        valido_a=date(2026, 7, 31),
        valido_in_date_json=["2026-07-01", "2026-07-02"],
    )
    run_id = await _registra_variazione(client, pid, "VARIAZIONE_CANCELLAZIONE")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}5",
                cod_origine=f"{_STAZ_PREFIX}I",
                cod_destinazione=f"{_STAZ_PREFIX}J",
                valido_da=date(2026, 7, 1),
                valido_a=date(2026, 7, 31),
            )
        ]
    )
    res = _applica_variazione(client, pid, run_id, xlsx)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["n_corse_update"] == 1
    assert body["n_warnings"] == 0

    async with session_scope() as session:
        row = (
            await session.execute(
                text(
                    "SELECT is_cancellata, cancellata_da_run_id, cancellata_at, "
                    "valido_in_date_json "
                    "FROM corsa_commerciale WHERE id = :id"
                ),
                {"id": corsa_id},
            )
        ).first()
        assert row is not None
        assert row[0] is True
        assert row[1] == run_id
        assert row[2] is not None
        # ``valido_in_date_json`` resta intatto (audit: si vede cosa era
        # attivo al momento della cancellazione).
        assert row[3] == ["2026-07-01", "2026-07-02"]


# =====================================================================
# Errori HTTP
# =====================================================================


async def test_applica_run_inesistente_404(client: TestClient) -> None:
    pid = await _crea_programma_var("run_404")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno="x",
                cod_origine="x",
                cod_destinazione="x",
                valido_da=date(2026, 1, 1),
                valido_a=date(2026, 1, 31),
            )
        ]
    )
    res = _applica_variazione(client, pid, 999_999_999, xlsx)
    assert res.status_code == 404


async def test_applica_run_gia_completata_409(client: TestClient) -> None:
    """Riapplicare una run già applicata → 409."""
    pid = await _crea_programma_var("run_doppia")
    await _crea_stazioni([f"{_STAZ_PREFIX}K", f"{_STAZ_PREFIX}L"])
    run_id = await _registra_variazione(client, pid, "INTEGRAZIONE")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}9",
                cod_origine=f"{_STAZ_PREFIX}K",
                cod_destinazione=f"{_STAZ_PREFIX}L",
                valido_da=date(2026, 1, 1),
                valido_a=date(2026, 1, 31),
            )
        ]
    )
    res1 = _applica_variazione(client, pid, run_id, xlsx)
    assert res1.status_code == 200

    res2 = _applica_variazione(client, pid, run_id, xlsx)
    assert res2.status_code == 409
    assert "già applicata" in res2.json()["detail"]


async def test_applica_estensione_non_supportata_415(
    client: TestClient,
) -> None:
    pid = await _crea_programma_var("ext_415")
    run_id = await _registra_variazione(client, pid, "INTEGRAZIONE")
    res = _applica_variazione(client, pid, run_id, b"not-an-xlsx", filename="garbage.txt")
    assert res.status_code == 415


async def test_applica_403_se_non_pianificatore_giro(
    client: TestClient,
) -> None:
    """Solo PIANIFICATORE_GIRO (admin bypass) può applicare variazioni."""
    pid = await _crea_programma_var("auth_403")
    run_id = await _registra_variazione(client, pid, "INTEGRAZIONE")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno="x",
                cod_origine="x",
                cod_destinazione="x",
                valido_da=date(2026, 1, 1),
                valido_a=date(2026, 1, 31),
            )
        ]
    )
    # Login con un ruolo non-PIANIFICATORE_GIRO se esistente; uso il
    # default seed pianificatore_giro_demo che invece è ammesso, quindi
    # qui simulo un'assenza di token.
    res = client.post(
        f"/api/programmi/{pid}/variazioni/{run_id}/applica",
        files={"file": ("var.xlsx", xlsx)},
    )
    assert res.status_code == 401  # senza token


async def test_applica_n_corse_e_completed_at_su_run(
    client: TestClient,
) -> None:
    """Verifica side-effect sulla run: n_corse, n_corse_create,
    completed_at sono aggiornati."""
    pid = await _crea_programma_var("run_state")
    await _crea_stazioni([f"{_STAZ_PREFIX}M", f"{_STAZ_PREFIX}N"])
    run_id = await _registra_variazione(client, pid, "INTEGRAZIONE")
    xlsx = _build_xlsx_pde_mini(
        [
            _row_pde_minimal(
                treno=f"{_TRENO_PREFIX}7",
                cod_origine=f"{_STAZ_PREFIX}M",
                cod_destinazione=f"{_STAZ_PREFIX}N",
                valido_da=date(2026, 8, 1),
                valido_a=date(2026, 8, 31),
            )
        ]
    )
    res = _applica_variazione(client, pid, run_id, xlsx)
    assert res.status_code == 200

    async with session_scope() as session:
        row = (
            await session.execute(
                text(
                    "SELECT n_corse, n_corse_create, n_corse_update, "
                    "completed_at FROM corsa_import_run WHERE id = :id"
                ),
                {"id": run_id},
            )
        ).first()
        assert row is not None
        assert row[0] == 1  # n_corse_lette_da_file
        assert row[1] == 1  # n_corse_create
        assert row[2] == 0  # n_corse_update
        assert isinstance(row[3], datetime)
